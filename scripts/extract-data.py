#!/usr/bin/env python3
"""Extract XLSX data into structured JSON files for the Astro site."""

import json
import re
import sys
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent.parent / "reference" / "em_director_matrix_v6.xlsx"
OUT_DIR = Path(__file__).parent.parent / "src" / "data"


def slugify(text: str) -> str:
    """Convert text to URL-friendly slug."""
    text = text.lower().strip()
    text = re.sub(r"[''']s\b", "s", text)  # possessives
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def read_sheet(wb, sheet_name):
    """Read a sheet into a list of dicts using the first row as headers."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        entry = {}
        for h, v in zip(headers, row):
            entry[h] = v
        data.append(entry)
    return data


def split_semicolons(val):
    """Split semicolon-delimited string into list, or return empty list."""
    if not val:
        return []
    return [x.strip() for x in str(val).split(";") if x.strip()]


def to_float(val, default=0.0):
    """Safely convert to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def to_int(val, default=0):
    """Safely convert to int."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_str(val, default=""):
    """Safely convert to string."""
    if val is None:
        return default
    return str(val).strip()


def extract_capabilities(raw):
    """Extract and structure capabilities."""
    caps = []
    for r in raw:
        cap_id = safe_str(r.get("CapabilityID"))
        name = safe_str(r.get("Name"))
        caps.append({
            "id": cap_id,
            "name": name,
            "slug": slugify(name),
            "domain": safe_str(r.get("Domain")),
            "description": safe_str(r.get("Description")),
            "primarySubTopics": split_semicolons(r.get("PrimarySubTopics")),
            "observableCount": to_int(r.get("ObservableCount")),
            "v5Status": safe_str(r.get("V5Status")),
        })
    return caps


def extract_observables(raw):
    """Extract and structure observables."""
    obs = []
    for r in raw:
        obs_id = safe_str(r.get("ObservableID"))
        short = safe_str(r.get("ShortText"))
        obs.append({
            "id": obs_id,
            "capabilityId": safe_str(r.get("CapabilityID")),
            "shortText": short,
            "slug": slugify(short) if short else slugify(obs_id),
            "fullExample": safe_str(r.get("FullExample")),
            "evidenceTypes": split_semicolons(r.get("EvidenceTypes")),
            "defaultWeight": to_float(r.get("DefaultWeight")),
            "requiredFrequency": safe_str(r.get("RequiredFrequency")),
            "emRelevance": safe_str(r.get("EMRelevance")),
            "directorRelevance": safe_str(r.get("DirectorRelevance")),
            "levelNotes": safe_str(r.get("LevelNotes")),
            "why": safe_str(r.get("Why")),
            "how": safe_str(r.get("How")),
            "expectedResult": safe_str(r.get("ExpectedResult")),
            "status": safe_str(r.get("Status")),
            "calibrationSignals": [],  # populated later
        })
    return obs


def extract_anti_patterns(raw):
    """Extract and structure anti-patterns."""
    aps = []
    for r in raw:
        ap_id = safe_str(r.get("AntiPatternID"))
        short = safe_str(r.get("ShortDesc"))
        # ShortDesc format: "Name — Description". Extract just the name for slug.
        name_part = short.split("—")[0].strip() if "—" in short else short
        aps.append({
            "id": ap_id,
            "name": name_part,
            "slug": slugify(name_part) if name_part else slugify(ap_id),
            "observableIds": split_semicolons(r.get("ObservableIDs")),
            "capabilityId": safe_str(r.get("CapabilityID")),
            "shortDesc": short,
            "warningSigns": safe_str(r.get("WarningSigns")),
            "impact": safe_str(r.get("Impact")),
            "recoveryActions": safe_str(r.get("RecoveryActions")),
            "sourceTopic": safe_str(r.get("SourceTopic")),
            "mappingNotes": safe_str(r.get("MappingNotes")),
        })
    return aps


def extract_calibration_signals(raw):
    """Extract and structure calibration signals."""
    sigs = []
    for r in raw:
        sig_id = safe_str(r.get("SignalID"))
        sigs.append({
            "id": sig_id,
            "observableId": safe_str(r.get("ObservableID")),
            "capabilityId": safe_str(r.get("CapabilityID")),
            "signalText": safe_str(r.get("SignalText")),
            "signalType": safe_str(r.get("SignalType")),
            "sourceSubTopic": safe_str(r.get("SourceSubTopic")),
        })
    return sigs


def extract_playbooks(raw):
    """Extract and structure playbooks."""
    pbs = []
    for r in raw:
        pb_id = safe_str(r.get("PlaybookID"))
        title = safe_str(r.get("Title"))
        pbs.append({
            "id": pb_id,
            "slug": slugify(title) if title else slugify(pb_id),
            "observableIds": split_semicolons(r.get("ObservableIDs")),
            "capabilityIds": split_semicolons(r.get("CapabilityIDs")),
            "title": title,
            "context": safe_str(r.get("Context")),
            "topicsActivated": split_semicolons(r.get("TopicsActivated")),
            "decisionFramework": safe_str(r.get("DecisionFramework")),
            "commonMistakes": safe_str(r.get("CommonMistakes")),
            "whatGoodLooksLike": safe_str(r.get("WhatGoodLooksLike")),
            "mappingNotes": safe_str(r.get("MappingNotes")),
        })
    return pbs


def extract_rubric_anchors(raw):
    """Extract and structure rubric anchors."""
    anchors = []
    for r in raw:
        anchors.append({
            "capabilityId": safe_str(r.get("CapabilityID")),
            "sourceTopic": safe_str(r.get("SourceTopic")),
            "level1Developing": safe_str(r.get("Level1_Developing")),
            "level3Competent": safe_str(r.get("Level3_Competent")),
            "level5Advanced": safe_str(r.get("Level5_Advanced")),
            "rationale": safe_str(r.get("Rationale")),
        })
    return anchors


def extract_evidence_types(raw):
    """Extract and structure evidence types."""
    evts = []
    for r in raw:
        key = safe_str(r.get("TypeKey"))
        evts.append({
            "typeKey": key,
            "slug": slugify(key),
            "displayName": safe_str(r.get("DisplayName")),
            "description": safe_str(r.get("Description")),
            "evidenceStrength": safe_str(r.get("EvidenceStrength")),
            "usageCount": to_int(r.get("UsageCount")),
        })
    return evts


def extract_crosswalk(raw):
    """Extract and structure crosswalk."""
    entries = []
    for r in raw:
        entries.append({
            "origTopic": safe_str(r.get("OrigTopic")),
            "origPrinciple": safe_str(r.get("OrigPrinciple")),
            "origSubTopic": safe_str(r.get("OrigSubTopic")),
            "primaryCapability": safe_str(r.get("PrimaryCapability")),
            "secondaryCapability": safe_str(r.get("SecondaryCapability")),
            "mappingNotes": safe_str(r.get("MappingNotes")),
            "observableId": safe_str(r.get("ObservableID")),
        })
    return entries


def extract_coverage(raw):
    """Extract and structure coverage data."""
    items = []
    for r in raw:
        items.append({
            "capabilityId": safe_str(r.get("CapabilityID")),
            "name": safe_str(r.get("Name")),
            "domain": safe_str(r.get("Domain")),
            "primarySubTopics": to_int(r.get("PrimarySubTopics")),
            "observables": to_int(r.get("Observables")),
            "coveragePercent": to_float(str(r.get("Coverage%", "0")).replace("%", "")),
        })
    return items


def nest_signals_into_observables(observables, signals):
    """Nest calibration signals into their parent observables."""
    sig_by_obs = {}
    for sig in signals:
        obs_id = sig["observableId"]
        sig_by_obs.setdefault(obs_id, []).append(sig)
    for obs in observables:
        obs["calibrationSignals"] = sig_by_obs.get(obs["id"], [])


def build_search_index(capabilities, observables, anti_patterns, playbooks):
    """Build a flat search index across all entity types."""
    index = []

    # Build capability lookup
    cap_map = {c["id"]: c for c in capabilities}

    for cap in capabilities:
        index.append({
            "title": f"{cap['id']}: {cap['name']}",
            "description": cap["description"],
            "type": "capability",
            "url": f"/capabilities/{cap['slug']}/",
            "domain": cap["domain"],
            "capabilityName": cap["name"],
            "id": cap["id"],
        })

    for obs in observables:
        cap = cap_map.get(obs["capabilityId"], {})
        index.append({
            "title": f"{obs['id']}: {obs['shortText']}",
            "description": obs["fullExample"],
            "type": "observable",
            "url": f"/capabilities/{cap.get('slug', '')}/#{obs['id'].lower()}",
            "domain": cap.get("domain", ""),
            "capabilityName": cap.get("name", ""),
            "id": obs["id"],
        })

    for ap in anti_patterns:
        cap = cap_map.get(ap["capabilityId"], {})
        index.append({
            "title": f"{ap['id']}: {ap['shortDesc']}",
            "description": ap["warningSigns"],
            "type": "anti-pattern",
            "url": f"/anti-patterns/{ap['slug']}/",
            "domain": cap.get("domain", ""),
            "capabilityName": cap.get("name", ""),
            "id": ap["id"],
        })

    for pb in playbooks:
        cap_names = [cap_map.get(cid, {}).get("name", "") for cid in pb["capabilityIds"]]
        domains = list(set(cap_map.get(cid, {}).get("domain", "") for cid in pb["capabilityIds"]))
        index.append({
            "title": f"{pb['id']}: {pb['title']}",
            "description": pb["context"],
            "type": "playbook",
            "url": f"/playbooks/{pb['slug']}/",
            "domain": domains[0] if domains else "",
            "capabilityName": ", ".join(cap_names),
            "id": pb["id"],
        })

    return index


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: XLSX file not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # Extract raw data from each sheet
    raw_caps = read_sheet(wb, "Capabilities")
    raw_obs = read_sheet(wb, "Observables")
    raw_aps = read_sheet(wb, "AntiPatterns")
    raw_sigs = read_sheet(wb, "CalibrationSignals")
    raw_pbs = read_sheet(wb, "Playbooks")
    raw_rubric = read_sheet(wb, "RubricAnchors")
    raw_evidence = read_sheet(wb, "EvidenceTypes_Ref")
    raw_crosswalk = read_sheet(wb, "Crosswalk")
    raw_coverage = read_sheet(wb, "Coverage_Plan")

    # Transform
    capabilities = extract_capabilities(raw_caps)
    observables = extract_observables(raw_obs)
    anti_patterns = extract_anti_patterns(raw_aps)
    signals = extract_calibration_signals(raw_sigs)
    playbooks = extract_playbooks(raw_pbs)
    rubric_anchors = extract_rubric_anchors(raw_rubric)
    evidence_types = extract_evidence_types(raw_evidence)
    crosswalk = extract_crosswalk(raw_crosswalk)
    coverage = extract_coverage(raw_coverage)

    # Nest signals into observables
    nest_signals_into_observables(observables, signals)

    # Build search index
    search_index = build_search_index(capabilities, observables, anti_patterns, playbooks)

    # Write JSON files
    outputs = {
        "capabilities.json": capabilities,
        "observables.json": observables,
        "anti-patterns.json": anti_patterns,
        "calibration-signals.json": signals,
        "playbooks.json": playbooks,
        "rubric-anchors.json": rubric_anchors,
        "evidence-types.json": evidence_types,
        "crosswalk.json": crosswalk,
        "coverage.json": coverage,
        "search-index.json": search_index,
    }

    for filename, data in outputs.items():
        path = OUT_DIR / filename
        with open(path, "w") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"  Wrote {filename}: {len(data)} entries")

    print(f"\nDone! {len(outputs)} files written to {OUT_DIR}")


if __name__ == "__main__":
    main()
